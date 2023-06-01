# -*- coding: utf-8 -*-
"""
Created on Sun May 28 22:06:08 2023

@author: H. Pfaff
"""
import openpyxl
import requests
import json
import re
import math

# Dateikonstanten
VERBRAUCHER = "Verbraucher.xlsx"
QUELLEN = "Quellen.xlsx"
ERGEBNIS = "Ergebnisse.xlsx"

# Kategorisierungskonstanten
KAT_5_MAX = 100
KAT_4_MAX = 200
KAT_3_MAX = 300
KAT_2_MAX = 400
KAT_1_MAX = 500

def adresskoordinaten_schreiben(worksheet, adressen):
    for i, adresse in enumerate(adressen, 1):
        ws.cell(row = i, column= 1, value = adresse)
        koordinaten = koordinaten_abrufen(i, adresse)
        ws.cell(row = i, column= 2, value = float(koordinaten[0]))
        ws.cell(row = i, column= 3, value = float(koordinaten[1]))

def adressen_abrufen(dateiname):
    workbook = openpyxl.load_workbook(dateiname)
    ws = workbook.active
    adress_liste = []
    for i, row in enumerate(ws.rows, 1):
        adress_liste.append(row[0].value)
    return adress_liste

def koordinaten_abrufen(i, adresse):
    split_adresse = re.split(',', adresse)
    plz_ort = split_adresse[-1]
    suchparameter = {"q": plz_ort}
    antwort = requests.get("https://geocode.maps.co/search", suchparameter)
    geodaten = json.loads(antwort.text)
    breite = geodaten[0]['lat']
    laenge = geodaten[0]['lon']
    return [breite, laenge]
    
def entfernung_berechnen(breite1, laenge1, breite2, laenge2):
    R = 6371
    dLat = math.radians(breite2 - breite1)
    dLon = math.radians(laenge2 - laenge1)
    a = math.sin(dLat/2) * math.sin(dLat/2) + \
        math.cos(math.radians(breite1)) * math.cos(math.radians(breite2)) * \
        math.sin(dLon/2) * math.sin(dLon/2)
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    d = R * c
    return d

def min_entfernungen_schreiben():
    for i, row in enumerate(wb["Verbraucher"].rows, 1):
        liste_entfernungen = []
        for j, row in enumerate(wb["Quellen"].rows, 1):
            entfernung = entfernung_berechnen(float(wb["Verbraucher"].cell(row = i, column= 2).value), 
                                              float(wb["Verbraucher"].cell(row = i, column= 3).value), 
                                              float(wb["Quellen"].cell(row = j, column= 2).value), 
                                              float(wb["Quellen"].cell(row = j, column= 3).value))
            liste_entfernungen.append(entfernung)
            min_entfernung = min(liste_entfernungen)
            min_index = [a for a, b in enumerate(liste_entfernungen) if b == min_entfernung]
        wb["Verbraucher"].cell(row = i, column= 4, value = min_entfernung)
        wb["Verbraucher"].cell(row = i, column= 5, value = "[" + str(min_index[0] + 1) + "]")
        wb["Verbraucher"].cell(row = i, column= 6, value = kategorisieren(min_entfernung))
        
def kategorisieren(entfernung):
    if entfernung < KAT_5_MAX:
        return 5
    elif entfernung < KAT_4_MAX:
        return 4
    elif entfernung < KAT_3_MAX:
        return 3
    elif entfernung < KAT_2_MAX:
        return 2
    elif entfernung < KAT_1_MAX:
        return 1
    else: 
        return 0
    
def formatieren_erweitert(blatt):
    formatieren(blatt)
    blatt.cell(row = 1, column= 5, value = "Entfernung in km")
    blatt.cell(row = 1, column= 6, value = "Quellenindex")
    blatt.cell(row = 1, column= 7, value = "Kategorie")
    
def formatieren(blatt):
    blatt.insert_rows(1,1)
    blatt.insert_cols(1,1)
    blatt.cell(row = 1, column= 1, value = "Index")
    blatt.cell(row = 1, column= 2, value = "Adresse")
    blatt.cell(row = 1, column= 3, value = "Breitengrad")
    blatt.cell(row = 1, column= 4, value = "Längengrad")
    for i, row in enumerate(blatt.rows, 1):
        blatt.cell(row = i + 1, column= 1, value = "[" + str(i) + "]")
    blatt.delete_rows(len(blatt['A']), 1)
    
wb = openpyxl.Workbook()
ws = wb.active

ws.title = "Verbraucher"
adresskoordinaten_schreiben(ws, adressen_abrufen(VERBRAUCHER))

wb.create_sheet("Quellen")
ws = wb["Quellen"]
adresskoordinaten_schreiben(ws, adressen_abrufen(QUELLEN))


min_entfernungen_schreiben()
formatieren_erweitert(wb["Verbraucher"])
formatieren(wb["Quellen"])

wb.save(ERGEBNIS)
